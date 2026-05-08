import pandas as pd
import json
import math
import subprocess
import os
import re
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.comments import Comment

# =========================
# CONFIGURAÇÕES
# =========================
ARQUIVO_EXCEL = r"\\SERV13-BKP\Serv13 Fazendas Arquivos & BKP\Fazenda Sweet Paraiso\13 - Planejamento Agrícola\Cronograma_agricola_Paraíso.xlsx"

PASTA_GITHUB = r"C:\Users\emanuel.rodrigues\Documents\GitHub\cronograma-paraiso"

ARQUIVO_JSON = os.path.join(PASTA_GITHUB, "dados.json")

ABA_ANO = "2026"

# =========================
# FUNÇÃO PARA LIMPAR COMENTÁRIO
# =========================
def limpar_comentario(texto):
    """
    Remove texto padrão do Excel como '[Comentário encadeado]' e links
    """
    if not texto:
        return texto
    
    # Remove o marcador de comentário encadeado
    texto = re.sub(r'\[Comentário encadeado\]\s*', '', texto)
    
    # Remove o aviso do Excel
    texto = re.sub(r'Sua versão do Excel permite que você leia este comentário encadeado.*?https?://[^\s]+', '', texto, flags=re.DOTALL)
    
    # Remove links restantes
    texto = re.sub(r'https?://[^\s]+', '', texto)
    
    # Remove linhas em branco extras
    texto = texto.strip()
    
    if not texto or texto == "":
        return None
    
    return texto

# =========================
# LER CABEÇALHO E MAPEAR COLUNAS
# =========================
print("📖 Lendo dados do Excel...")

# Primeiro, lê o cabeçalho para saber os nomes das colunas
df_raw = pd.read_excel(
    ARQUIVO_EXCEL,
    sheet_name=ABA_ANO,
    header=None
)

linha_cabecalho = None

for i, row in df_raw.iterrows():
    if row.astype(str).str.lower().str.contains("parcela").any():
        linha_cabecalho = i
        break

if linha_cabecalho is None:
    raise Exception("Cabeçalho com PARCELA não encontrado.")

print(f"📌 Cabeçalho encontrado na linha {linha_cabecalho + 1}")

# Lê o cabeçalho para mapear colunas
df_header = pd.read_excel(
    ARQUIVO_EXCEL,
    sheet_name=ABA_ANO,
    header=linha_cabecalho,
    nrows=0  # Lê apenas o cabeçalho
)

# Cria um mapeamento de número da coluna (1-based) para nome da coluna
colunas = list(df_header.columns)
coluna_mapeamento = {}
for idx, col in enumerate(colunas):
    # Excel é 1-based, então coluna = idx + 1
    coluna_mapeamento[idx + 1] = str(col).strip()

print(f"📊 {len(colunas)} colunas encontradas")

# =========================
# LER COMENTÁRIOS DO EXCEL
# =========================
print("\n📝 Lendo comentários do Excel...")

def ler_comentarios_excel_com_precisao(arquivo_excel, aba_nome, coluna_mapeamento, linha_cabecalho):
    """
    Lê comentários e mapeia para o campo correto usando o cabeçalho
    """
    comentarios = {}
    
    try:
        workbook = load_workbook(arquivo_excel, data_only=True)
        
        if aba_nome in workbook.sheetnames:
            planilha = workbook[aba_nome]
            
            for row in range(1, planilha.max_row + 1):
                for col in range(1, planilha.max_column + 1):
                    cell = planilha.cell(row=row, column=col)
                    
                    if cell.comment and isinstance(cell.comment, Comment):
                        texto_original = cell.comment.text
                        texto_limpo = limpar_comentario(texto_original)
                        
                        if texto_limpo:
                            # Verifica se é linha de dados (após cabeçalho)
                            if row > linha_cabecalho + 1:
                                # Obtém o nome da coluna pelo mapeamento
                                nome_coluna = coluna_mapeamento.get(col, f"Coluna_{col}")
                                
                                # Obtém o número da parcela nesta linha
                                # Procura a coluna que contém "PARCELA"
                                parcela_col = None
                                for c_idx, c_nome in coluna_mapeamento.items():
                                    if "parcela" in c_nome.lower():
                                        parcela_col = c_idx
                                        break
                                
                                if parcela_col:
                                    cell_parcela = planilha.cell(row=row, column=parcela_col)
                                    parcela_num = cell_parcela.value
                                    
                                    if parcela_num and pd.notna(parcela_num):
                                        try:
                                            parcela_num = int(parcela_num)
                                            key = (parcela_num, nome_coluna)
                                            comentarios[key] = texto_limpo
                                            print(f"📝 Parcela {parcela_num}, Coluna '{nome_coluna}': {texto_limpo[:50]}...")
                                        except:
                                            pass
        else:
            print(f"⚠️ Aba '{aba_nome}' não encontrada")
        
        workbook.close()
        
    except Exception as e:
        print(f"⚠️ Erro ao ler comentários: {e}")
    
    return comentarios

comentarios = ler_comentarios_excel_com_precisao(
    ARQUIVO_EXCEL, 
    ABA_ANO, 
    coluna_mapeamento, 
    linha_cabecalho
)

print(f"📊 Total de comentários encontrados: {len(comentarios)}")

# =========================
# CARREGAR DADOS COMPLETOS
# =========================
df = pd.read_excel(
    ARQUIVO_EXCEL,
    sheet_name=ABA_ANO,
    header=linha_cabecalho
)

# =========================
# LIMPAR COLUNAS
# =========================
df.columns = df.columns.astype(str).str.strip()
df = df.loc[:, ~df.columns.str.contains("Unnamed", case=False)]

# =========================
# FORÇAR PARCELA INTEIRO
# =========================
coluna_parcela = next(
    (c for c in df.columns if "parcela" in c.lower()),
    None
)

if coluna_parcela:
    df[coluna_parcela] = pd.to_numeric(
        df[coluna_parcela],
        errors="coerce"
    ).fillna(0).astype(int)

# =========================
# CONVERTER DADOS
# =========================
def converter(valor):
    if pd.isna(valor):
        return None

    if isinstance(valor, int):
        return int(valor)

    if isinstance(valor, float):
        if math.isnan(valor):
            return None

        if valor.is_integer():
            return int(valor)

        return valor

    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, str):
        v = valor.strip()

        if v == "" or v == "-":
            return None

        return v

    return valor

df = df.apply(lambda col: col.apply(converter))

# =========================
# ADICIONAR COMENTÁRIOS AOS DADOS
# =========================
print("\n🔗 Vinculando comentários às parcelas...")

# Converte o DataFrame para lista de dicionários
dados = df.to_dict(orient="records")

# Adiciona comentários a cada parcela
for idx, linha in enumerate(dados):
    parcela_num = linha.get(coluna_parcela)
    comentarios_parcela = {}
    
    # Busca comentários para esta parcela
    for (parcela, coluna), texto in comentarios.items():
        if parcela == parcela_num:
            comentarios_parcela[coluna] = texto
    
    if comentarios_parcela:
        linha['COMENTARIOS'] = comentarios_parcela
        print(f"✅ Parcela {parcela_num}: {len(comentarios_parcela)} comentário(s)")

# =========================
# LIMPAR NAN FINAL
# =========================
def limpar_nan(obj):
    if isinstance(obj, float) and math.isnan(obj):
        return None

    if isinstance(obj, dict):
        return {k: limpar_nan(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [limpar_nan(v) for v in obj]

    return obj

dados = limpar_nan(dados)

# =========================
# SALVAR JSON
# =========================
with open(ARQUIVO_JSON, "w", encoding="utf-8") as f:
    json.dump(
        dados,
        f,
        ensure_ascii=False,
        indent=2,
        allow_nan=False
    )

print(f"\n✅ JSON atualizado com {len(dados)} registros")

# =========================
# ENVIAR GITHUB
# =========================
os.chdir(PASTA_GITHUB)

subprocess.run(["git", "add", "."], check=True)

status = subprocess.run(
    ["git", "status", "--porcelain"],
    capture_output=True,
    text=True
)

if status.stdout.strip():
    subprocess.run(
        ["git", "commit", "-m", "Atualização automática com comentários"],
        check=True
    )

    subprocess.run(
        ["git", "push"],
        check=True
    )

    print("✅ GitHub atualizado com sucesso.")

else:
    print("ℹ️ Nenhuma alteração detectada. Nada para enviar.")